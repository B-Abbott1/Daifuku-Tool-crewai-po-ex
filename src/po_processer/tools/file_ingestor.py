from pathlib import Path


def ingest_file(file_path: str) -> str:
    """Read a PO file (PDF, Excel, or plain text) and return structured text."""
    path = Path(file_path)
    if not path.exists():
        return f"[FileIngestorTool] File not found: {file_path}"

    suffix = path.suffix.lower()

    if suffix == ".pdf":
        return _ingest_pdf(path)
    if suffix in (".xlsx", ".xls"):
        return _ingest_excel(path)
    if suffix in (".txt", ".csv"):
        return _ingest_text(path)
    return f"[FileIngestorTool] Unsupported file type: {suffix}"


def _ingest_pdf(path: Path) -> str:
    try:
        import pdfplumber
    except ImportError:
        return "[FileIngestorTool] pdfplumber is not installed. Run: pip install pdfplumber"

    output_parts = []

    with pdfplumber.open(str(path)) as pdf:
        for page_num, page in enumerate(pdf.pages, start=1):
            output_parts.append(f"--- Page {page_num} ---")

            tables = page.extract_tables()

            if tables:
                for table_idx, table in enumerate(tables, start=1):
                    if len(tables) > 1:
                        output_parts.append(f"[Table {table_idx}]")
                    output_parts.append(_format_table(table))
            else:
                text = page.extract_text(x_tolerance=3, y_tolerance=3)
                if text:
                    output_parts.append(text.strip())

    return "\n".join(output_parts)


def _format_table(table: list) -> str:
    rows = []
    for row in table:
        cleaned = [
            (cell.strip().replace("\n", " ") if cell else "")
            for cell in row
        ]
        rows.append("\t".join(cleaned))
    return "\n".join(rows)


def _ingest_excel(path: Path) -> str:
    try:
        import openpyxl
    except ImportError:
        return "[FileIngestorTool] openpyxl is not installed. Run: pip install openpyxl"

    wb = openpyxl.load_workbook(str(path), data_only=True)
    output_parts = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        output_parts.append(f"--- Sheet: {sheet_name} ---")
        for row in ws.iter_rows(values_only=True):
            cleaned = [str(cell) if cell is not None else "" for cell in row]
            output_parts.append("\t".join(cleaned))

    return "\n".join(output_parts)


def _ingest_text(path: Path) -> str:
    try:
        return path.read_text(encoding="utf-8")
    except UnicodeDecodeError:
        return path.read_text(encoding="latin-1")
