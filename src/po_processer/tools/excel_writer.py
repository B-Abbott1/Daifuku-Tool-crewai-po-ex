import json

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

DARK_BLUE = "1F3864"
WHITE = "FFFFFF"


def _extract_first_json_object(text: str) -> dict:
    text = text.strip()
    if text.startswith("```"):
        text = text.split("```")[1]
        if text.startswith("json"):
            text = text[4:]
        text = text.strip()

    decoder = json.JSONDecoder()
    obj, _ = decoder.raw_decode(text)
    return obj


def write_excel(normalized_data: str, output_path: str = "po_output.xlsx") -> str:
    try:
        po_data = _extract_first_json_object(normalized_data)
    except Exception as e:
        return f"[ExcelWriterTool] Failed to parse input JSON: {e}"

    try:
        wb = Workbook()
        _write_line_items_sheet(wb, po_data)

        if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
            del wb["Sheet"]

        wb.save(output_path)

        line_count = len(po_data.get("line_items", []))
        return (
            f"Excel workbook written to '{output_path}' — "
            f"Sheet: 'Line Items' — {line_count} line item(s) written."
        )
    except Exception as e:
        return f"[ExcelWriterTool] Failed to write Excel file: {e}"


def _apply_header(cell, value: str) -> None:
    cell.value = value
    cell.font = Font(name="Arial", bold=True, size=11, color=WHITE)
    cell.fill = PatternFill("solid", fgColor=DARK_BLUE)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _autosize_columns(ws, min_width: int = 12, max_width: int = 60) -> None:
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            try:
                max_len = max(max_len, len(str(cell.value or "")))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = max(min_width, min(max_len + 2, max_width))


def _write_line_items_sheet(wb: Workbook, po_data: dict) -> None:
    ws = wb.active
    ws.title = "Line Items"

    columns = [
        "line_number",
        "part_number",
        "description",
        "quantity",
        "unit_of_measure",
        "unit_price",
        "total_price",
        "supplier_part_number",
        "notes",
    ]

    for col_idx, col_name in enumerate(columns, start=1):
        _apply_header(
            ws.cell(row=1, column=col_idx),
            col_name.replace("_", " ").title(),
        )

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}1"

    for row_idx, item in enumerate(po_data.get("line_items", []), start=2):
        row_data = [
            item.get("line_number"),
            item.get("part_number"),
            item.get("description"),
            item.get("quantity"),
            item.get("unit_of_measure"),
            item.get("unit_price"),
            item.get("total_price"),
            item.get("supplier_part_number"),
            item.get("notes"),
        ]
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = Font(name="Arial", size=10)
            cell.alignment = Alignment(wrap_text=False, vertical="center")

    _autosize_columns(ws)
